"""Generate payroll_test_errors.xlsx with intentional errors for testing the auditing app."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Payroll"

# Headers
headers = [
    "Employee_id", "first_name", "last_name", "gender", "department",
    "hourly_rate", "hours_worked", "gross_pay", "tax_deduction", "net_pay",
]
header_font = Font(bold=True)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font

# ---------- normal rows ----------
normal_employees = [
    (1, "Anne",   "Nelson",   "Female", "Engineering",          28, 40, 1120.00, 112.00,  1008.00),
    (2, "Rose",   "Griffin",  "Female", "Product Development",  22, 38,  836.00,  83.60,   752.40),
    (3, "James",  "Carter",   "Male",   "Operations",           30, 42, 1260.00, 126.00,  1134.00),
    (4, "Julia",  "Flores",   "Female", "Product Development",  25, 40, 1000.00, 100.00,   900.00),
    (5, "Mark",   "Adams",    "Male",   "Customer Support",     20, 36,  720.00,  72.00,   648.00),
]

# ---------- error rows ----------
error_rows = [
    # Error 2: duplicate of employee 1
    (1, "Anne",   "Nelson",   "Female", "Engineering",          28, 40, 1120.00, 112.00,  1008.00),
    # Error 3: negative gross pay
    (6, "Tom",    "Baker",    "Male",   "Marketing",            18, 35, -630.00,  63.00,  -693.00),
    # Error 4: outlier hourly rate ($1200/hr)
    (7, "Sarah",  "Lee",      "Female", "Engineering",        1200, 40, 48000.00, 4800.00, 43200.00),
    # Error 5: excessive hours (168 in a week)
    (8, "Kevin",  "Moore",    "Male",   "Operations",           22, 168, 3696.00, 369.60,  3326.40),
    # Error 6: tax rate > 50% (tax is 60% of gross)
    (9, "Diana",  "Price",    "Female", "Human Resources",      26, 40, 1040.00, 624.00,   416.00),
    # Error 7: net pay != gross - deductions (off by $200)
    (10, "Brian", "Hall",     "Male",   "Sales",                24, 40,  960.00,  96.00,   664.00),
]

row_num = 2
all_gross = []

for emp in normal_employees + error_rows:
    for col, val in enumerate(emp, 1):
        ws.cell(row=row_num, column=col, value=val)
    all_gross.append(emp[7])  # gross_pay
    row_num += 1

# Error 8: missing values row
ws.cell(row=row_num, column=1, value=11)
ws.cell(row=row_num, column=2, value="Pat")
# last_name, gender, department, hourly_rate left blank
ws.cell(row=row_num, column=8, value=800.00)
ws.cell(row=row_num, column=9, value=80.00)
ws.cell(row=row_num, column=10, value=720.00)
all_gross.append(800.00)
row_num += 1

# Error 9: #REF! error string in a cell
ws.cell(row=row_num, column=1, value=12)
ws.cell(row=row_num, column=2, value="Error")
ws.cell(row=row_num, column=3, value="Row")
ws.cell(row=row_num, column=4, value="Male")
ws.cell(row=row_num, column=5, value="Engineering")
ws.cell(row=row_num, column=6, value=25)
ws.cell(row=row_num, column=7, value=40)
ws.cell(row=row_num, column=8, value="#REF!")  # spreadsheet error string
ws.cell(row=row_num, column=9, value=100.00)
ws.cell(row=row_num, column=10, value=900.00)
row_num += 1

# Error 1: Total row with wrong sum (off by $5000)
actual_sum = sum(all_gross)
ws.cell(row=row_num, column=1, value="TOTAL")
ws.cell(row=row_num, column=8, value=actual_sum + 5000)
ws.cell(row=row_num, column=8).font = Font(bold=True)

# Auto-size columns
for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        for cell in row:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 2

# Format currency columns
for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=8, max_col=10):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

wb.save("payroll_test_errors.xlsx")
print("Created payroll_test_errors.xlsx")
print(f"  {len(normal_employees)} normal rows")
print(f"  {len(error_rows)} error rows")
print(f"  + 1 missing-values row")
print(f"  + 1 #REF! error row")
print(f"  + 1 total row (wrong sum: {actual_sum + 5000} vs actual {actual_sum})")
